#!/usr/bin/env python3
"""
Gera planilha Excel template para importar dados no app MakeLifeBetter.
Uso: python3 gerar_planilha_template.py
"""

import subprocess
import sys

# Instalar openpyxl se necessario
try:
    import openpyxl
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl"])
    import openpyxl

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = Workbook()

# ========================================
# ABA 1: Eventos
# ========================================
ws_eventos = wb.active
ws_eventos.title = "Eventos"

# Estilos
header_font = Font(bold=True, color="FFFFFF", size=12)
header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
header_alignment = Alignment(horizontal="center", vertical="center")
thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin")
)

# Cabecalhos
headers_eventos = ["titulo", "subtitulo", "descricao", "hora", "lugar", "categoria"]
for col, header in enumerate(headers_eventos, 1):
    cell = ws_eventos.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = header_alignment
    cell.border = thin_border

# Dados de exemplo
eventos = [
    ["Cerimonia de Abertura", "Bem-vindos ao evento", "Cerimonia oficial de abertura com discursos e apresentacoes especiais.", "09:00", "Salao Principal", "cerimonia"],
    ["Coffee Break", "Pausa para cafe", "Momento de networking e degustacao de cafe e lanches.", "10:30", "Area de Convivencia", "intervalo"],
    ["Palestra Principal", "Tema especial do dia", "Palestra inspiradora sobre inovacao e tecnologia.", "11:00", "Auditorio", "palestra"],
    ["Almoco", "Refeicao", "Almoco servido no restaurante do local.", "12:30", "Restaurante", "refeicao"],
    ["Workshop Pratico", "Atividade pratica", "Workshop interativo com atividades em grupo.", "14:00", "Sala de Treinamento", "workshop"],
    ["Encerramento", "Despedida", "Cerimonia de encerramento e agradecimentos.", "17:00", "Salao Principal", "cerimonia"],
]

for row_idx, evento in enumerate(eventos, 2):
    for col_idx, value in enumerate(evento, 1):
        cell = ws_eventos.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

# Ajustar largura das colunas
ws_eventos.column_dimensions["A"].width = 30
ws_eventos.column_dimensions["B"].width = 25
ws_eventos.column_dimensions["C"].width = 60
ws_eventos.column_dimensions["D"].width = 15
ws_eventos.column_dimensions["E"].width = 25
ws_eventos.column_dimensions["F"].width = 15

# ========================================
# ABA 2: Localizacao
# ========================================
ws_local = wb.create_sheet("Localizacao")

headers_local = ["name", "address", "city", "latitude", "longitude"]
for col, header in enumerate(headers_local, 1):
    cell = ws_local.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    cell.alignment = header_alignment
    cell.border = thin_border

# Dados de exemplo
local_data = ["Centro de Convencoes", "Av. Principal, 1000", "Sao Paulo - SP", -23.550520, -46.633308]
for col_idx, value in enumerate(local_data, 1):
    cell = ws_local.cell(row=2, column=col_idx, value=value)
    cell.border = thin_border

ws_local.column_dimensions["A"].width = 30
ws_local.column_dimensions["B"].width = 30
ws_local.column_dimensions["C"].width = 25
ws_local.column_dimensions["D"].width = 15
ws_local.column_dimensions["E"].width = 15

# ========================================
# ABA 3: Contatos
# ========================================
ws_contatos = wb.create_sheet("Contatos")

headers_contatos = ["name", "phone"]
for col, header in enumerate(headers_contatos, 1):
    cell = ws_contatos.cell(row=1, column=col, value=header)
    cell.font = header_font
    cell.fill = PatternFill(start_color="BF8F00", end_color="BF8F00", fill_type="solid")
    cell.alignment = header_alignment
    cell.border = thin_border

# Dados de exemplo
contatos = [
    ["Recepcao", "(11) 1234-5678"],
    ["Organizacao", "(11) 9876-5432"],
    ["Emergencia", "(11) 9999-9999"],
]

for row_idx, contato in enumerate(contatos, 2):
    for col_idx, value in enumerate(contato, 1):
        cell = ws_contatos.cell(row=row_idx, column=col_idx, value=value)
        cell.border = thin_border

ws_contatos.column_dimensions["A"].width = 25
ws_contatos.column_dimensions["B"].width = 20

# ========================================
# Salvar
# ========================================
output_path = "template_importacao.xlsx"
wb.save(output_path)
print(f"Planilha template criada com sucesso: {output_path}")
print()
print("A planilha possui 3 abas:")
print("  1. Eventos    - titulo, subtitulo, descricao, hora, lugar, categoria")
print("  2. Localizacao - name, address, city, latitude, longitude")
print("  3. Contatos   - name, phone")
print()
print("Edite os dados de exemplo e importe no app!")
